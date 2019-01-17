import random
import os
import time
import shutil

from PyQt5.QtCore import *
import openpyxl
from openpyxl.styles import Side, Border


class CheckPoint(QObject):
    def __init__(self, point_name, check_time, check_water, check_date = None):
        super().__init__()

        self.point_name = point_name
        self.check_time = check_time
        self.check_water = check_water
        self.check_date = time.strftime("%Y/%m/%d %H:%M:%S", time.localtime())[:11] if check_date is None else check_date

        self.flow_list = []
        self.water_list = []
        self.time_list = []

    def get_flow_excel(self):
        flow0 = self.check_water / self.check_time / 1000

        """流量信息"""
        flow00 = flow0
        flow01 = flow0 * random.uniform(0.95, 1.05)
        flow02 = flow0 * random.uniform(0.95, 1.05)
        flow10 = flow0 * random.uniform(1.25, 1.35)
        flow11 = flow0 * random.uniform(1.25, 1.35)
        flow12 = flow0 * random.uniform(1.25, 1.35)
        flow20 = flow0 * random.uniform(1.28, 1.37)
        flow21 = flow0 * random.uniform(1.28, 1.37)
        flow22 = flow0 * random.uniform(1.28, 1.37)
        self.flow_list=[flow00, flow01, flow02,flow10, flow11, flow12,flow20, flow21, flow22]
        # for i in self.flow_list:
        #     print('flow:{}'.format(round(i, 3)))

        """单次流量"""
        water00 = self.check_water
        water01 = random.randrange(1500, 3500, 100)
        water02 = random.randrange(1500, 3500, 100)
        water10 = random.randrange(1500, 3500, 100)
        water11 = random.randrange(1500, 3500, 100)
        water12 = random.randrange(1500, 3500, 100)
        water20 = random.randrange(1500, 3500, 100)
        water21 = random.randrange(1500, 3500, 100)
        water22 = random.randrange(1500, 3500, 100)
        self.water_list = [water00,water01,water02,water10,water11,water12,water20,water21,water22]
        # for i in self.water_list:
        #     print('water:{}'.format(i))

        """单次时间"""
        time00 = self.check_time
        time01 = water01 / (flow01 * 1000)
        time02 = water02 / (flow02 * 1000)
        time10 = water10 / (flow10 * 1000)
        time11 = water11 / (flow11 * 1000)
        time12 = water12 / (flow12 * 1000)
        time20 = water20 / (flow20 * 1000)
        time21 = water21 / (flow21 * 1000)
        time22 = water22 / (flow22 * 1000)
        self.time_list = [time00,time01,time02,time10,time11,time12,time20,time21,time22]
        # for i in self.time_list:
        #     print('time:{}'.format(round(i, 2)))

        """填写流量表信息"""
        project_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
        orogin_excel = os.path.abspath(os.path.join(project_path, 'resource/容器法外业记录表.xlsx'))
        output_folder = os.path.abspath(os.path.join(project_path, 'output'))
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)
        wb = openpyxl.load_workbook(orogin_excel)
        ws = wb['流量']
        ws['A4'].value = self.point_name
        for i in range(4, 13):
            temp_time = 'D'+str(i)
            ws[temp_time].value = round(self.time_list[i-4], 2)
            temp_water = 'E'+str(i)
            ws[temp_water].value = self.water_list[i-4]

        """填写日期信息"""
        print(ws['A2'].value[5:15])
        til_str = ws['A2'].value
        date_str = ws['A2'].value[5:15]
        ws['A2'].value = til_str.replace(date_str, self.check_date)

        """修复表边框"""
        left, right, top, bottom = [Side(style='thin', color='000000')] * 4
        border = Border(left=left, right=right, top=top, bottom=bottom)
        for row in ws.rows:
            for cell in row:
                cell.border = border

        """保存表文件"""
        point_folder = os.path.abspath(os.path.join(output_folder, self.point_name))
        if os.path.exists(point_folder):
            wb.save(os.path.abspath(os.path.join(point_folder, self.point_name + '.xlsx')))
        else:
            os.makedirs(point_folder)
            wb.save(os.path.abspath(os.path.join(point_folder, self.point_name + '.xlsx')))

    def get_sample_excel(self):
        pass


if __name__ == '__main__':
    path = '/media/sheildog/183E3B373E3B0D6E/Workspace/12-24到12-25'
    for folder in os.listdir(path):
        folder_path = os.path.abspath(os.path.join(path, folder))
        if os.path.isdir(folder_path):
            picture_folder = os.path.abspath(os.path.join(folder_path, '照片'))
            if not os.path.exists(picture_folder):
                os.makedirs(picture_folder)
        for file in os.listdir(folder_path):
            (file_name, file_extension) = os.path.splitext(file)
            if file_extension == '.jpg':
                picture_path = os.path.abspath(os.path.join(folder_path, file))
                p_folder = os.path.abspath(os.path.join(folder_path, '照片'))
                shutil.move(picture_path, p_folder)
                print(file)
